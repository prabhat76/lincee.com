#!/usr/bin/env python3
"""
Generate comprehensive API documentation Excel file with all endpoints,
request/response examples, and error handling
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API Documentation"

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
section_font = Font(bold=True, size=11)
error_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
error_font = Font(color="9C6500", size=10)
success_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
success_font = Font(color="375623", size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 50
ws.column_dimensions['F'].width = 50
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 40

row = 1

# Title
ws['A1'] = 'LINCEE E-COMMERCE API DOCUMENTATION'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:H1')
row = 3

# Headers
headers = ['API #', 'Endpoint', 'Method', 'URI', 'Request Body', 'Response (200)', 'Status Code', 'Error Handling']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

row += 1

# API Endpoints Data
apis = [
    # AUTH ENDPOINTS
    {
        'section': 'AUTHENTICATION',
        'endpoint': 'User Login',
        'method': 'POST',
        'uri': '/api/v1/auth/login',
        'request': '{"username": "admin", "password": "password123"}',
        'response': '{"token": "eyJhbGc...", "userId": 1, "role": "ADMIN"}',
        'codes': '200, 400, 401',
        'errors': '400: Invalid credentials | 401: Unauthorized'
    },
    {
        'endpoint': 'User Register',
        'method': 'POST',
        'uri': '/api/v1/auth/register',
        'request': '{"username": "john_doe", "email": "john@example.com", "password": "pass123"}',
        'response': '{"id": 5, "username": "john_doe", "email": "john@example.com", "role": "CUSTOMER"}',
        'codes': '201, 400, 409',
        'errors': '400: Invalid data | 409: Email/username already exists'
    },
    {
        'endpoint': 'User Logout',
        'method': 'POST',
        'uri': '/api/v1/auth/logout',
        'request': 'Authorization: Bearer <token>',
        'response': '{"message": "Logout successful"}',
        'codes': '200, 401',
        'errors': '401: Invalid token'
    },
    {
        'endpoint': 'Refresh Token',
        'method': 'POST',
        'uri': '/api/v1/auth/refresh',
        'request': '{"refreshToken": "eyJhbGc..."}',
        'response': '{"token": "new_jwt_token"}',
        'codes': '200, 401',
        'errors': '401: Invalid or expired refresh token'
    },

    # PRODUCT ENDPOINTS
    {
        'section': 'PRODUCTS',
        'endpoint': 'Get All Products',
        'method': 'GET',
        'uri': '/api/v1/products?size=100&page=0',
        'request': 'No body required',
        'response': '{"content": [{...}, {...}], "totalElements": 52, "totalPages": 1}',
        'codes': '200, 400',
        'errors': '400: Invalid pagination parameters'
    },
    {
        'endpoint': 'Get Product by ID',
        'method': 'GET',
        'uri': '/api/v1/products/{id}',
        'request': 'No body required',
        'response': '{"id": 1, "name": "Classic Hoodie", "price": 49.99, "stock": 50}',
        'codes': '200, 404',
        'errors': '404: Product not found'
    },
    {
        'endpoint': 'Create Product',
        'method': 'POST',
        'uri': '/api/v1/products',
        'request': '{"name": "New Hoodie", "price": 59.99, "category": "Hoodies", "brand": "Lincee"}',
        'response': '{"id": 53, "name": "New Hoodie", ...}',
        'codes': '201, 400, 401',
        'errors': '400: Validation failed | 401: Unauthorized (Admin required)'
    },
    {
        'endpoint': 'Update Product',
        'method': 'PUT',
        'uri': '/api/v1/products/{id}',
        'request': '{"name": "Updated Hoodie", "price": 64.99}',
        'response': '{"id": 1, "name": "Updated Hoodie", "price": 64.99}',
        'codes': '200, 400, 404, 401',
        'errors': '401: Unauthorized | 404: Product not found | 400: Validation failed'
    },
    {
        'endpoint': 'Delete Product',
        'method': 'DELETE',
        'uri': '/api/v1/products/{id}',
        'request': 'No body required',
        'response': 'HTTP 204 No Content',
        'codes': '204, 404, 401',
        'errors': '401: Unauthorized | 404: Product not found'
    },
    {
        'endpoint': 'Search Products',
        'method': 'GET',
        'uri': '/api/v1/products/search?keyword=hoodie&size=20',
        'request': 'No body required',
        'response': '{"content": [{...}], "totalElements": 12}',
        'codes': '200, 400',
        'errors': '400: Invalid search parameters'
    },
    {
        'endpoint': 'Get by Category',
        'method': 'GET',
        'uri': '/api/v1/products/category/Hoodies?size=20',
        'request': 'No body required',
        'response': '{"content": [{...}], "totalElements": 12}',
        'codes': '200, 400',
        'errors': '400: Invalid category'
    },
    {
        'endpoint': 'Get Featured Products',
        'method': 'GET',
        'uri': '/api/v1/products/featured',
        'request': 'No body required',
        'response': '[{"id": 1, "name": "Hoodie 1", "featured": true}, ...]',
        'codes': '200',
        'errors': 'None'
    },

    # ORDER ENDPOINTS
    {
        'section': 'ORDERS',
        'endpoint': 'Create Order',
        'method': 'POST',
        'uri': '/api/v1/orders?userId=4',
        'request': '{"totalAmount": 169.96, "shippingAddressId": 1, "billingAddressId": 1, "orderItems": [...]}',
        'response': '{"id": 101, "orderNumber": "ORD-1234...", "status": "PENDING", "totalAmount": 169.96}',
        'codes': '201, 400, 404',
        'errors': '400: Invalid order data | 404: User/Product not found'
    },
    {
        'endpoint': 'Get Order by ID',
        'method': 'GET',
        'uri': '/api/v1/orders/{id}',
        'request': 'No body required',
        'response': '{"id": 101, "orderNumber": "ORD-1234", "status": "PENDING"}',
        'codes': '200, 404',
        'errors': '404: Order not found'
    },
    {
        'endpoint': 'Get User Orders',
        'method': 'GET',
        'uri': '/api/v1/orders/user/{userId}?page=0&size=10',
        'request': 'No body required',
        'response': '{"content": [{...}], "totalElements": 5}',
        'codes': '200, 404',
        'errors': '404: User not found'
    },
    {
        'endpoint': 'Update Order Status',
        'method': 'PATCH',
        'uri': '/api/v1/orders/{id}/status?status=SHIPPED',
        'request': 'No body required',
        'response': '{"id": 101, "status": "SHIPPED"}',
        'codes': '200, 400, 404',
        'errors': '400: Invalid status | 404: Order not found'
    },
    {
        'endpoint': 'Delete Order',
        'method': 'DELETE',
        'uri': '/api/v1/orders/{id}',
        'request': 'No body required',
        'response': 'HTTP 204 No Content',
        'codes': '204, 404, 401',
        'errors': '401: Unauthorized | 404: Order not found'
    },

    # CART ENDPOINTS
    {
        'section': 'SHOPPING CART',
        'endpoint': 'Get Cart',
        'method': 'GET',
        'uri': '/api/v1/cart/user/{userId}',
        'request': 'No body required',
        'response': '{"id": 1, "userId": 4, "items": [...], "totalPrice": 199.98}',
        'codes': '200, 404',
        'errors': '404: Cart not found'
    },
    {
        'endpoint': 'Add to Cart',
        'method': 'POST',
        'uri': '/api/v1/cart/user/{userId}/items',
        'request': '{"productId": 10, "quantity": 2, "size": "M", "color": "Black"}',
        'response': '{"id": 1, "cartItems": [...]}',
        'codes': '201, 400, 404',
        'errors': '400: Invalid item data | 404: Product/User not found'
    },
    {
        'endpoint': 'Update Cart Item',
        'method': 'PUT',
        'uri': '/api/v1/cart/items/{cartItemId}',
        'request': '{"quantity": 3}',
        'response': '{"id": 25, "quantity": 3}',
        'codes': '200, 400, 404',
        'errors': '400: Invalid quantity | 404: Cart item not found'
    },
    {
        'endpoint': 'Remove from Cart',
        'method': 'DELETE',
        'uri': '/api/v1/cart/user/{userId}/items/{cartItemId}',
        'request': 'No body required',
        'response': 'HTTP 204 No Content',
        'codes': '204, 404',
        'errors': '404: Cart item not found'
    },
    {
        'endpoint': 'Clear Cart',
        'method': 'DELETE',
        'uri': '/api/v1/cart/user/{userId}/clear',
        'request': 'No body required',
        'response': '{"message": "Cart cleared"}',
        'codes': '200, 404',
        'errors': '404: Cart not found'
    },

    # USER ENDPOINTS
    {
        'section': 'USERS',
        'endpoint': 'Get All Users',
        'method': 'GET',
        'uri': '/api/v1/users?page=0&size=20',
        'request': 'No body required',
        'response': '{"content": [{...}], "totalElements": 50}',
        'codes': '200, 401',
        'errors': '401: Unauthorized (Admin required)'
    },
    {
        'endpoint': 'Get User by ID',
        'method': 'GET',
        'uri': '/api/v1/users/{id}',
        'request': 'No body required',
        'response': '{"id": 4, "username": "john_doe", "email": "john@example.com"}',
        'codes': '200, 404, 401',
        'errors': '401: Unauthorized | 404: User not found'
    },
    {
        'endpoint': 'Update User',
        'method': 'PUT',
        'uri': '/api/v1/users/{id}',
        'request': '{"firstName": "John", "lastName": "Doe", "phone": "+919876543210"}',
        'response': '{"id": 4, "firstName": "John", "lastName": "Doe"}',
        'codes': '200, 400, 404, 401',
        'errors': '400: Validation failed | 404: User not found | 401: Unauthorized'
    },

    # ADDRESS ENDPOINTS
    {
        'section': 'ADDRESSES',
        'endpoint': 'Get User Addresses',
        'method': 'GET',
        'uri': '/api/v1/addresses/user/{userId}',
        'request': 'No body required',
        'response': '[{"id": 1, "street": "123 Main St", "city": "Hazaribagh"}]',
        'codes': '200, 404',
        'errors': '404: User not found'
    },
    {
        'endpoint': 'Create Address',
        'method': 'POST',
        'uri': '/api/v1/addresses/user/{userId}',
        'request': '{"street": "456 Oak Ave", "city": "Nawabganj", "state": "Bihar", "zip": "825301"}',
        'response': '{"id": 2, "userId": 4, "street": "456 Oak Ave"}',
        'codes': '201, 400, 404',
        'errors': '400: Validation failed | 404: User not found'
    },
    {
        'endpoint': 'Update Address',
        'method': 'PUT',
        'uri': '/api/v1/addresses/{id}',
        'request': '{"street": "789 Elm St", "city": "Patna"}',
        'response': '{"id": 1, "street": "789 Elm St", "city": "Patna"}',
        'codes': '200, 400, 404',
        'errors': '400: Validation failed | 404: Address not found'
    },
    {
        'endpoint': 'Delete Address',
        'method': 'DELETE',
        'uri': '/api/v1/addresses/{id}',
        'request': 'No body required',
        'response': 'HTTP 204 No Content',
        'codes': '204, 404',
        'errors': '404: Address not found'
    },

    # PAYMENT ENDPOINTS
    {
        'section': 'PAYMENTS',
        'endpoint': 'Create Payment Intent',
        'method': 'POST',
        'uri': '/api/v1/payments/intent',
        'request': '{"orderId": 101, "amount": 169.96, "currency": "usd"}',
        'response': '{"clientSecret": "pi_1234...", "paymentIntentId": "pi_1234"}',
        'codes': '201, 400, 404',
        'errors': '400: Invalid amount | 404: Order not found'
    },
    {
        'endpoint': 'Get Payment Status',
        'method': 'GET',
        'uri': '/api/v1/payments/{paymentId}',
        'request': 'No body required',
        'response': '{"id": 50, "orderId": 101, "status": "succeeded", "amount": 169.96}',
        'codes': '200, 404',
        'errors': '404: Payment not found'
    },

    # REVIEW ENDPOINTS
    {
        'section': 'REVIEWS',
        'endpoint': 'Create Review',
        'method': 'POST',
        'uri': '/api/v1/reviews?productId=10&userId=4',
        'request': '{"rating": 5, "title": "Great product!", "comment": "Excellent quality"}',
        'response': '{"id": 15, "productId": 10, "userId": 4, "rating": 5}',
        'codes': '201, 400, 404',
        'errors': '400: Invalid review data | 404: Product/User not found'
    },
    {
        'endpoint': 'Get Product Reviews',
        'method': 'GET',
        'uri': '/api/v1/reviews/product/{productId}?page=0&size=10',
        'request': 'No body required',
        'response': '{"content": [{...}], "totalElements": 25}',
        'codes': '200, 404',
        'errors': '404: Product not found'
    },
    {
        'endpoint': 'Update Review',
        'method': 'PUT',
        'uri': '/api/v1/reviews/{id}',
        'request': '{"rating": 4, "comment": "Good product"}',
        'response': '{"id": 15, "rating": 4, "comment": "Good product"}',
        'codes': '200, 400, 404',
        'errors': '400: Validation failed | 404: Review not found'
    },
    {
        'endpoint': 'Delete Review',
        'method': 'DELETE',
        'uri': '/api/v1/reviews/{id}',
        'request': 'No body required',
        'response': 'HTTP 204 No Content',
        'codes': '204, 404',
        'errors': '404: Review not found'
    },

    # COLLECTION ENDPOINTS
    {
        'section': 'COLLECTIONS',
        'endpoint': 'Get Collections',
        'method': 'GET',
        'uri': '/api/v1/collections',
        'request': 'No body required',
        'response': '[{"id": 1, "name": "Summer Collection", "productCount": 15}]',
        'codes': '200',
        'errors': 'None'
    },
    {
        'endpoint': 'Get Collection by ID',
        'method': 'GET',
        'uri': '/api/v1/collections/{id}',
        'request': 'No body required',
        'response': '{"id": 1, "name": "Summer Collection", "products": [...]}',
        'codes': '200, 404',
        'errors': '404: Collection not found'
    },
    {
        'endpoint': 'Create Collection',
        'method': 'POST',
        'uri': '/api/v1/collections/admin',
        'request': '{"name": "Winter Sale", "description": "Winter products on sale"}',
        'response': '{"id": 5, "name": "Winter Sale"}',
        'codes': '201, 400, 401',
        'errors': '400: Validation failed | 401: Unauthorized (Admin required)'
    },

    # DASHBOARD ENDPOINTS
    {
        'section': 'DASHBOARD',
        'endpoint': 'Get Dashboard Stats',
        'method': 'GET',
        'uri': '/api/v1/dashboard/stats',
        'request': 'No body required',
        'response': '{"totalOrders": 150, "totalRevenue": 25000, "totalUsers": 45}',
        'codes': '200, 401',
        'errors': '401: Unauthorized (Admin required)'
    },

    # HEALTH CHECK
    {
        'section': 'HEALTH CHECK',
        'endpoint': 'Health Status',
        'method': 'GET',
        'uri': '/health',
        'request': 'No body required',
        'response': '{"status": "UP", "timestamp": "2026-02-14T10:00:00"}',
        'codes': '200',
        'errors': 'None'
    },
]

# Write API data
api_count = 0
for api in apis:
    if 'section' in api:
        # Write section header
        ws[f'A{row}'] = api['section']
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.fill = section_fill
        cell.font = section_font
        cell.border = border
        row += 1
    else:
        # Write API row
        api_count += 1
        ws[f'A{row}'] = api_count
        ws[f'B{row}'] = api.get('endpoint', '')
        ws[f'C{row}'] = api.get('method', '')
        ws[f'D{row}'] = api.get('uri', '')
        ws[f'E{row}'] = api.get('request', '')
        ws[f'F{row}'] = api.get('response', '')
        ws[f'G{row}'] = api.get('codes', '')
        ws[f'H{row}'] = api.get('errors', '')

        # Apply borders and wrapping
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            if col == 8:  # Error column
                cell.fill = error_fill
                cell.font = error_font
            elif api.get('codes', '').startswith('200') or api.get('codes', '') == 'None':
                cell.fill = success_fill
                cell.font = success_font

        row += 1

# Add error code reference sheet
ws2 = wb.create_sheet("HTTP Status Codes & Error Enums")
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 50

row = 1
ws2[f'A{row}'] = 'HTTP Status Codes'
ws2[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
ws2[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws2.merge_cells(f'A{row}:C{row}')
row += 1

status_codes = [
    ('200', 'OK', 'Request successful'),
    ('201', 'Created', 'Resource created successfully'),
    ('204', 'No Content', 'Request successful, no content to return'),
    ('400', 'Bad Request', 'Validation failed or missing required fields'),
    ('401', 'Unauthorized', 'Authentication required or token invalid'),
    ('403', 'Forbidden', 'User does not have permission'),
    ('404', 'Not Found', 'Resource not found'),
    ('409', 'Conflict', 'Duplicate resource (e.g., email already exists)'),
    ('422', 'Unprocessable Entity', 'Validation error on input'),
    ('500', 'Internal Server Error', 'Server error, try again later'),
    ('503', 'Service Unavailable', 'Server temporarily unavailable'),
]

for code, name, desc in status_codes:
    ws2[f'A{row}'] = code
    ws2[f'B{row}'] = name
    ws2[f'C{row}'] = desc
    for col in range(1, 4):
        cell = ws2.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# Add error enum reference
row += 2
ws2[f'A{row}'] = 'Common Error Response Enums'
ws2[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
ws2[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws2.merge_cells(f'A{row}:C{row}')
row += 1

error_enums = [
    ('VALIDATION_ERROR', '400', '{"code": "VALIDATION_ERROR", "message": "Field required", "field": "email"}'),
    ('UNAUTHORIZED', '401', '{"code": "UNAUTHORIZED", "message": "Invalid credentials"}'),
    ('FORBIDDEN', '403', '{"code": "FORBIDDEN", "message": "Only admin can access"}'),
    ('RESOURCE_NOT_FOUND', '404', '{"code": "RESOURCE_NOT_FOUND", "message": "Product not found"}'),
    ('DUPLICATE_RESOURCE', '409', '{"code": "DUPLICATE_RESOURCE", "message": "Email already registered"}'),
    ('INSUFFICIENT_STOCK', '400', '{"code": "INSUFFICIENT_STOCK", "message": "Only 5 items in stock"}'),
    ('INVALID_ORDER', '400', '{"code": "INVALID_ORDER", "message": "Order contains no items"}'),
    ('PAYMENT_FAILED', '400', '{"code": "PAYMENT_FAILED", "message": "Payment declined"}'),
    ('INTERNAL_ERROR', '500', '{"code": "INTERNAL_ERROR", "message": "An unexpected error occurred"}'),
]

for error, code, response in error_enums:
    ws2[f'A{row}'] = error
    ws2[f'B{row}'] = code
    ws2[f'C{row}'] = response
    for col in range(1, 4):
        cell = ws2.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if col == 1:
            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    row += 1

# Save workbook
wb.save('/Users/prabhatkumar/Documents/lincee-backend/LINCEE_API_DOCUMENTATION.xlsx')
print("✅ API Documentation Excel file created: LINCEE_API_DOCUMENTATION.xlsx")
print(f"📊 Total endpoints documented: {api_count}")
